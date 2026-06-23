"""
EDC Parameter Config — Excel to JSON Converter
วิธีใช้:
  1. วางไฟล์ Excel ทุกแบงค์ไว้ในโฟลเดอร์เดียวกับ script นี้
  2. รัน: python convert.py
  3. จะได้ data.json สำหรับ deploy

รองรับ: KBANK, SCB, BBL, BAY (CID/CID3/CID4)
"""

import json
import os
import glob
import openpyxl
from datetime import datetime

OUTPUT_FILE = "data.json"

# ──────────────────────────────────────────
# Config แต่ละแบงค์ — ปรับ column name ให้ตรงกับไฟล์จริง
# ──────────────────────────────────────────
BANK_CONFIGS = {
    "KBANK": {
        "file_pattern": "*KBANK*.xlsx",
        "sheet": None,          # None = sheet แรก
        "tid_col": "TID",
        "mid_col": "MID",
        "sn_col": "S/N EDC",
        "slip_col": "Slip Line_1",
        "merchant_col": "Slip Line_1",
        "sw_col": "SW Version",
        "model_col": "Model",
        # function columns (Y/N หรือ 1/0)
        "func_cols": {
            "KBANK_PRE_AUTH": "FUNCTION_KBANK_PRE_AUTH",
            "KBANK_TIP": "FUNCTION_KBANK_TIP",
            "KBANK_OFFLINE": "FUNCTION_KBANK_OFFLINE",
            "KBANK_KEYIN": "FUNCTION_KBANK_KEYIN",
            "KBANK_REFUND": "FUNCTION_KBANK_REFUND",
            "TPN_PRE_AUTH": "FUNCTION_TPN_PRE_AUTH",
            "TPN_PRE_AUTH_CANCEL": "FUNCTION_TPN_PRE_AUTH_CANCEL",
            "TPN_SALE_COMPLETE_ONLINE": "FUNCTION_TPN_SALE_COMPLETE_ONLINE",
            "TPN_REFUND": "FUNCTION_TPN_REFUND",
            "TPN_KEYIN": "FUNCTION_TPN_KEYIN",
        },
        # sub TID columns (แสดงเพิ่มเติม)
        "sub_tid_cols": ["TID_DCC", "TID_SmartPay", "TID_TPN", "TID_Alipay",
                         "TID_Wechat", "TID_BBL_AMEX", "TID_Redemption",
                         "TID_QR_Credit_Card", "TID_K_CHECK_ID"],
    },
    "SCB": {
        "file_pattern": "*SCB*.xlsx",
        "sheet": None,
        "tid_col": "HOST.1/TERMINAL_ID",
        "mid_col": "HOST.1/MERCHANT_ID",
        "sn_col": "SERIAL_NUMBER",
        "slip_col": "MERCHANT_NAME_LINE1",
        "merchant_col": "MERCHANT_NAME_LINE1",
        "sw_col": "SW_VERSION",
        "model_col": "MODEL",
        "func_cols": {},
        "sub_tid_cols": [],
    },
    "WOM": {
        "file_pattern": "*WOM*.xlsx",
        "sheet": None,          # None = sheet แรก
        "tid_col": "MAIN_TID",
        "mid_col": "MAIN_MID",
        "sn_col": "SERIAL_NO",
        "slip_col": "SLIP_LINE_1",
	    "slip_col": "SLIP_LINE_2",
	    "slip_col": "SLIP_LINE_3",
        "merchant_col": "Slip Line_1",
        "sw_col": "SW_VERSION",
        "model_col": "Model",
        # function columns (Y/N หรือ 1/0 หรือ '1/'0)
        "func_cols": {
            "KBANK_PRE_AUTH": "FUNCTION_KBANK_PRE_AUTH",
            "KBANK_TIP": "FUNCTION_KBANK_TIP",
            "KBANK_OFFLINE": "FUNCTION_KBANK_OFFLINE",
            "FUNCTION_KBANK_KEYIN": "FUNCTION_KBANK_KEYIN",
            "KBANK_REFUND": "FUNCTION_KBANK_REFUND",
            "TPN_PRE_AUTH": "FUNCTION_TPN_PRE_AUTH",
            "TPN_PRE_AUTH_CANCEL": "FUNCTION_TPN_PRE_AUTH_CANCEL",
            "TPN_SALE_COMPLETE_ONLINE": "FUNCTION_TPN_SALE_COMPLETE_ONLINE",
            "TPN_REFUND": "FUNCTION_TPN_REFUND",
            "TPN_KEYIN": "FUNCTION_TPN_KEYIN",
        },
        # sub TID columns (แสดงเพิ่มเติม)
        "sub_tid_cols": ["TID_DCC", "TID_SMARTPAY", "TID_TPN", "TID_Alipay",
                         "TID_Wechat", "TID_BBL_AMEX", "TID_REDEMPTION",
                         "TID_QR_Credit_Card", "TID_KPLUS", "TID_K_CHECK_ID"],
    },
    "BBL": {
        "file_pattern": "*BBL*.xlsx",
        "sheet": None,
        "tid_col": "Terminal ID",
        "mid_col": "MID",
        "sn_col": "Serial No.",
        "slip_col": "LINE 1",
        "merchant_col": "LINE 1",
        "sw_col": "SW Version",
        "model_col": "Model",
        "func_cols": {},
        "sub_tid_cols": ["TID UPI"],
    },
    "BAY": {
        "file_pattern": "*BAY*.xlsx",
        "sheet": None,
        "tid_col": "TIDหลัก",
        "mid_col": "MID BAY",
        "sn_col": None,
        "slip_col": "TM LINE1",
        "merchant_col": "TM LINE1",
        "sw_col": None,
        "model_col": None,
        "func_cols": {
            "B25": "B25",
            "B21": "B21",
            "B27": "B27",
        },
        "sub_tid_cols": ["TID DCC", "TID QRCS", "TID QRALIPAY",
                         "TID QRPROMPT", "TID KSC", "TID BBL", "TID KBANK"],
    },
}

# ──────────────────────────────────────────
def col_index(headers, name):
    """หา index ของ column จากชื่อ (case-insensitive)"""
    if name is None:
        return None
    nl = name.lower().strip()
    for i, h in enumerate(headers):
        if h and str(h).lower().strip() == nl:
            return i
    return None

def safe(val):
    """แปลงค่าให้ปลอดภัยสำหรับ JSON"""
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("None", "NULL", "nan", "0.0"):
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s

def is_enabled(val):
    """เช็คว่า function เปิดอยู่ไหม"""
    if val is None:
        return False
    s = str(val).upper().strip()
    return s in ("Y", "YES", "1", "TRUE", "ENABLED", "✓")

# ──────────────────────────────────────────
def process_bank(bank_name, config):
    records = []
    files = glob.glob(config["file_pattern"])
    if not files:
        print(f"  [{bank_name}] ไม่พบไฟล์ pattern: {config['file_pattern']}")
        return records

    for filepath in files:
        print(f"  [{bank_name}] กำลังอ่าน: {filepath}")
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb[config["sheet"]] if config["sheet"] else wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            tid_i      = col_index(headers, config["tid_col"])
            mid_i      = col_index(headers, config["mid_col"])
            sn_i       = col_index(headers, config["sn_col"])
            slip_i     = col_index(headers, config["slip_col"])
            sw_i       = col_index(headers, config["sw_col"])
            model_i    = col_index(headers, config["model_col"])

            func_indices = {
                k: col_index(headers, v)
                for k, v in config["func_cols"].items()
            }
            sub_tid_indices = {
                name: col_index(headers, name)
                for name in config["sub_tid_cols"]
            }

            if tid_i is None:
                print(f"    ⚠️  ไม่พบ column TID '{config['tid_col']}' — ข้ามไฟล์นี้")
                continue

            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                tid_raw = row[tid_i] if tid_i is not None else None
                if not tid_raw:
                    continue
                tid = safe(tid_raw)
                if not tid or tid in ("", "0"):
                    continue

                # functions ที่เปิด
                funcs = [k for k, i in func_indices.items()
                         if i is not None and is_enabled(row[i])]

                # sub TIDs
                sub_tids = {
                    name: safe(row[i])
                    for name, i in sub_tid_indices.items()
                    if i is not None and safe(row[i])
                }

                rec = {
                    "bank":     bank_name,
                    "tid":      tid,
                    "mid":      safe(row[mid_i]) if mid_i is not None else "",
                    "sn":       safe(row[sn_i])  if sn_i  is not None else "",
                    "merchant": safe(row[slip_i]) if slip_i is not None else "",
                    "sw":       safe(row[sw_i])   if sw_i   is not None else "",
                    "model":    safe(row[model_i]) if model_i is not None else "",
                    "funcs":    funcs,
                    "sub_tids": sub_tids,
                }
                records.append(rec)
                row_count += 1

            print(f"    ✓ {row_count:,} records")
            wb.close()

        except Exception as e:
            print(f"    ❌ Error: {e}")

    return records

# ──────────────────────────────────────────
def main():
    print("=" * 50)
    print("EDC Parameter Config — Excel to JSON Converter")
    print("=" * 50)

    all_records = []
    for bank_name, config in BANK_CONFIGS.items():
        records = process_bank(bank_name, config)
        all_records.extend(records)

    output = {
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "total": len(all_records),
        "records": all_records,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_FILE) / 1024
    print()
    print(f"✅ เสร็จแล้ว! {len(all_records):,} records → {OUTPUT_FILE} ({size_kb:.0f} KB)")
    print(f"   อัพเดทล่าสุด: {output['updated']}")

if __name__ == "__main__":
    main()
